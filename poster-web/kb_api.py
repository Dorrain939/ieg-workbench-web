"""知识库 API：文档上传、列表、删除、检索。

数据目录：kb_data/
  ├── global/<doc_id>/
  └── projects/<pid>/<doc_id>/
  └── functions/<pid>/<function_id>/<kb_type>/<doc_id>/
每个 doc_id 目录下：
  - meta.json      文档元信息
  - original.<ext> 原文件
  - text.txt       解析后的纯文本
  - chunks.json    切片结果（LLM/检索用）
"""
import json
import pathlib
import shutil
import time
import uuid
import datetime
from typing import Optional, List

from fastapi import APIRouter, UploadFile, File, Form, HTTPException
from fastapi.responses import FileResponse, JSONResponse

from kb import KB_DATA_DIR
from kb.loader import load as load_doc, LoaderError, SUPPORTED_EXTS, IMAGE_EXTS, TABLE_EXTS
from kb.chunker import chunk as chunk_text
from kb import index as kb_index


router = APIRouter(prefix="/api/kb")


# 单文件上传大小上限 20 MB
MAX_FILE_SIZE = 20 * 1024 * 1024


M_KEYWORDS = {
    "module.m1_text": ["项目背景", "项目简介", "致谢", "概述", "背景"],
    "module.m2_highlight_text": ["注意事项", "报名要求", "温馨提示", "下期预告", "核心问题", "悬念", "须知"],
    "module.m3_text_subsections": ["小标题", "Q&A", "问答", "提纲", "多段", "导师寄语", "分享内容"],
    "module.m4_plain_text": ["tips", "小tips", "小提示", "祝福语"],
    "module.m5_text_table": ["时间安排", "日程", "课程表", "安排表", "矩阵", "表格", "时间", "地点"],
    "module.m6_image_text_single": ["学习地图", "项目全景", "单图", "上图下文", "左图右文"],
    "module.m7_image_text_subsections": ["学员之声", "反馈分类", "图文反馈", "收获", "建议", "亮点"],
    "module.m8_single_image_text": ["二维码", "扫码", "单张图片", "报名码"],
    "module.m9_multi_image_text": ["奖品展示", "多图", "展示图"],
    "module.m10_single_person_card": ["嘉宾详介", "单个讲师", "主讲人", "讲师简介", "研究方向", "经历背景"],
    "module.m11_person_cards_row": ["几位讲师", "多张头像卡片"],
    "module.m12_avatar_wall": ["讲师阵容", "讲师团", "嘉宾阵容", "头像墙"],
    "module.m13_avatar_wall_groups": ["分组讲师", "多门课讲师", "分组头像", "父子头像"],
    "module.m14_text_name_list": ["学员名单", "表彰名单", "优秀学员", "大量学员"],
    "module.m15_text_name_list_groups": ["分组名单", "部门名单", "班级名单"],
    "module.m16_course_speaker_split": ["左讲师", "右课程", "课程卡片"],
    "module.m17_course_text_speaker": ["直播预告", "上文字", "下讲师"],
    "module.m18_course_parent_children": ["多门课程", "课程卡片", "逐课", "课程反馈卡片"],
    "module.m19_rating_bars": ["评分", "满意度", "分数", "课程评分"],
    "module.m20_single_image": ["大合影", "单张大图", "海报图"],
    "module.m21_multi_image_collage": ["照片墙", "活动剪影", "精彩瞬间", "作品展示", "成果展示", "多图拼盘"],
    "module.m22_button_inside": ["模块内按钮"],
    "module.m23_button_outside": ["报名按钮", "回放入口", "CTA", "立即报名", "按钮"],
    "module.m24_contact_text": ["联系人", "联系方式", "咨询"],
    "module.m25_contact_qr": ["联系人二维码", "二维码", "扫码联系"],
}


def _guess_type_scene(text: str) -> tuple[str, str]:
    lower = text.lower()
    if any(k in text for k in ["系列课程", "课程矩阵", "逐课", "每门课", "按课报名"]):
        return ("C", "S6b" if any(k in text for k in ["反馈", "评分", "结束后", "回顾"]) else "S6a")
    if any(k in text for k in ["主题分享", "分享会", "嘉宾", "单场", "回放"]):
        return ("B", "S5b" if any(k in text for k in ["回顾", "反馈", "回放", "结束后"]) else "S5a")
    if any(k in text for k in ["结项", "全面回顾", "毕业", "答辩"]):
        return ("A", "S3")
    if any(k in text for k in ["阶段总结", "阶段回顾", "第", "集中回顾"]):
        return ("A", "S2")
    if "成果展示" in text:
        return ("A", "S4")
    if any(k in text for k in ["开班", "招募", "报名", "启动", "开营"]):
        return ("A", "S1")
    return ("A", "S1")


def _infer_module_keys(text: str) -> list[str]:
    found = []
    explicit = []
    for i in range(1, 26):
        if f"M{i}" in text or f"m{i}" in text:
            explicit.append(i)
    if explicit:
        by_num = {int(k.split(".m", 1)[1].split("_", 1)[0]): k for k in M_KEYWORDS}
        found.extend(by_num[i] for i in explicit if i in by_num)
    for key, words in M_KEYWORDS.items():
        if any(w in text for w in words) and key not in found:
            found.append(key)
    return found


def _module_key_from_code(code: str) -> Optional[str]:
    raw = (code or "").strip().upper()
    if not raw.startswith("M"):
        return None
    digits = ""
    for ch in raw[1:]:
        if ch.isdigit():
            digits += ch
        else:
            break
    if not digits:
        return None
    prefix = f"module.m{int(digits)}_"
    for key in M_KEYWORDS:
        if key.startswith(prefix):
            return key
    return None


def _strategy_from_module_rows(project_type: str, scene: str, title: str, module_rows: list[dict]) -> dict:
    import poster_strategy
    import poster_module_registry

    strategy = poster_strategy.resolve_strategy(project_type, scene)
    registry = poster_module_registry.list_registry()["module_capabilities"]

    visual = [m for m in strategy.get("module_plan", []) if m.get("script_key") == "module.tm1_tm13_visual_layer"]
    modules = list(visual)
    used = {}
    for row in module_rows:
        key = _module_key_from_code(row.get("code", ""))
        if not key or key not in registry:
            continue
        used[key] = used.get(key, 0) + 1
        cap = registry[key]
        code = row.get("code", "").strip().upper()
        module_title = (row.get("title") or cap.get("label") or code).strip()
        modules.append({
            "id": f"{code}-KB-{used[key]}",
            "name": module_title,
            "display_name": cap.get("label") or module_title,
            "purpose": "由项目知识库表格指定",
            "required": True,
            "component": cap.get("renderer") or "spec_text_panel",
            "script_key": key,
            "status": cap.get("status") or "script_pending",
            "status_label": cap.get("status_label"),
            "input_schema_key": f"input.{key}",
            "qa_rules_key": f"qa.{key}",
            "module_config": {
                "module_title": module_title,
                "title_enabled": True,
                "content": "",
                "images": [],
            },
        })
    strategy["module_plan"] = modules
    strategy["required_count"] = sum(1 for m in strategy.get("module_plan", []) if m.get("required"))
    strategy["optional_count"] = sum(1 for m in strategy.get("module_plan", []) if not m.get("required"))
    enriched = poster_module_registry.enrich_strategy(strategy)
    enriched["recognition"] = {
        "project_type": project_type,
        "scene": scene,
        "matched_modules": [m.get("script_key") for m in modules],
        "source": "project_kb_table",
        "title": title,
    }
    return enriched


def _strategy_from_text(text: str) -> dict:
    import poster_strategy
    import poster_module_registry

    project_type, scene = _guess_type_scene(text)
    strategy = poster_strategy.resolve_strategy(project_type, scene)
    inferred_keys = _infer_module_keys(text)
    existing = [m.get("script_key") for m in strategy.get("module_plan", [])]
    registry = poster_module_registry.list_registry()["module_capabilities"]
    for key in inferred_keys:
        if key in existing or key not in registry:
            continue
        cap = registry[key]
        code = key.split(".", 1)[1].split("_", 1)[0].upper()
        strategy["module_plan"].append({
            "id": f"{code}-KB",
            "name": cap.get("label") or code,
            "purpose": "由项目知识库识别补充",
            "required": False,
            "component": cap.get("renderer") or "spec_text_panel",
            "script_key": key,
            "status": "script_pending",
            "input_schema_key": f"input.{key}",
            "qa_rules_key": f"qa.{key}",
        })
    strategy["required_count"] = sum(1 for m in strategy.get("module_plan", []) if m.get("required"))
    strategy["optional_count"] = sum(1 for m in strategy.get("module_plan", []) if not m.get("required"))
    enriched = poster_module_registry.enrich_strategy(strategy)
    enriched["recognition"] = {
        "project_type": project_type,
        "scene": scene,
        "matched_modules": inferred_keys,
        "source": "project_kb_rules",
    }
    return enriched


def _original_file_for_meta(meta: dict) -> Optional[pathlib.Path]:
    ddir = _doc_dir(
        meta.get("scope") or "global",
        meta.get("project_id"),
        meta.get("id"),
        function_id=meta.get("function_id"),
        kb_type=meta.get("kb_type"),
    )
    matches = list(ddir.glob("original.*"))
    return matches[0] if matches else None


def _cell_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _recognize_excel_function_projects(path: pathlib.Path) -> list[dict]:
    if path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return []
    try:
        from openpyxl import load_workbook
    except Exception:
        return []

    groups = []
    try:
        wb = load_workbook(path, data_only=True)
    except Exception:
        return []

    for ws in wb.worksheets:
        rows = [[_cell_text(c) for c in row] for row in ws.iter_rows(values_only=True)]
        if not rows:
            continue
        header_idx = None
        col = {}
        for idx, row in enumerate(rows[:20]):
            normalized = "".join(row)
            if "项目类型" in normalized and ("海报子类型" in normalized or "场景" in normalized):
                header_idx = idx
                for ci, name in enumerate(row):
                    if "项目类型" in name:
                        col["project_type"] = ci
                    elif "海报子类型" in name or "场景" in name:
                        col["scene"] = ci
                    elif "项目名称" in name:
                        col["project_name"] = ci
                    elif "时间节点" in name or "海报名称" in name or "海报标题" in name:
                        col["poster_name"] = ci
                    elif "模块标题" in name:
                        col["module_title"] = ci
                    elif "模块类型编号" in name or "模块编号" in name:
                        col["module_code"] = ci
                break
        if header_idx is None or not {"project_type", "scene", "module_title", "module_code"}.issubset(col):
            continue

        current = None
        for row in rows[header_idx + 1:]:
            project_type = row[col["project_type"]] if col["project_type"] < len(row) else ""
            scene = row[col["scene"]] if col["scene"] < len(row) else ""
            project_name = row[col.get("project_name", 999)] if col.get("project_name", 999) < len(row) else ""
            poster_name = row[col.get("poster_name", 999)] if col.get("poster_name", 999) < len(row) else ""
            module_title = row[col["module_title"]] if col["module_title"] < len(row) else ""
            module_code = row[col["module_code"]] if col["module_code"] < len(row) else ""

            if project_type or scene or project_name or poster_name:
                if current and current["modules"]:
                    groups.append(current)
                current = {
                    "project_type": project_type or "A",
                    "scene": scene or "S1",
                    "project_name": project_name,
                    "poster_name": poster_name,
                    "modules": [],
                    "sheet": ws.title,
                }
            if current and (module_title or module_code):
                current["modules"].append({"title": module_title or module_code, "code": module_code})
        if current and current["modules"]:
            groups.append(current)

    projects = []
    for idx, group in enumerate(groups, start=1):
        poster_name = group.get("poster_name") or f"海报子项目 {idx}"
        project_name = group.get("project_name") or ""
        title = f"{project_name}｜{poster_name}" if project_name else poster_name
        strategy = _strategy_from_module_rows(
            group.get("project_type") or "A",
            group.get("scene") or "S1",
            title,
            group.get("modules") or [],
        )
        projects.append({
            "name": title,
            "description": f"{group.get('project_type') or 'A'}-{group.get('scene') or 'S1'} · 来自 {path.name}",
            "project_type": strategy.get("project_type", {}).get("id") or group.get("project_type") or "A",
            "scene": strategy.get("scene", {}).get("id") or group.get("scene") or "S1",
            "poster_strategy": strategy,
            "source": {
                "filename": path.name,
                "sheet": group.get("sheet"),
                "poster_name": poster_name,
            },
        })
    return projects


def _recognition_payload_from_sources(text: str, files: list[pathlib.Path]) -> dict:
    function_projects = []
    for fp in files:
        function_projects.extend(_recognize_excel_function_projects(fp))
    if function_projects:
        primary = function_projects[0]["poster_strategy"]
        payload = json.loads(json.dumps(primary, ensure_ascii=False))
        payload["function_projects"] = function_projects
        payload["recognition"]["function_project_count"] = len(function_projects)
        return payload
    primary = _strategy_from_text(text)
    primary["function_projects"] = [{
        "name": primary.get("recognition", {}).get("title") or primary.get("scene", {}).get("label") or "海报项目",
        "description": primary.get("scene", {}).get("goal") or "",
        "project_type": primary.get("project_type", {}).get("id"),
        "scene": primary.get("scene", {}).get("id"),
        "poster_strategy": primary,
        "source": {"filename": "text"},
    }]
    return primary


def _doc_text(meta: dict) -> str:
    ddir = _doc_dir(
        meta.get("scope") or "global",
        meta.get("project_id"),
        meta.get("id"),
        function_id=meta.get("function_id"),
        kb_type=meta.get("kb_type"),
    )
    fp = ddir / "text.txt"
    if not fp.exists():
        return ""
    try:
        return fp.read_text(encoding="utf-8")
    except Exception:
        return ""


def _now_iso() -> str:
    return datetime.datetime.now().isoformat(timespec="seconds")


def _gen_doc_id() -> str:
    return f"doc_{uuid.uuid4().hex[:12]}"


def _content_kind(ext: str) -> str:
    if ext in IMAGE_EXTS:
        return "image"
    if ext in TABLE_EXTS:
        return "table"
    return "text"


def _safe_part(value: Optional[str], name: str) -> str:
    value = (value or "").strip()
    if not value or "/" in value or ".." in value:
        raise HTTPException(400, f"非法 {name}: {value}")
    return value


def _scope_dir(
    scope: str,
    project_id: Optional[str],
    function_id: Optional[str] = None,
    kb_type: Optional[str] = None,
) -> pathlib.Path:
    """返回 scope 对应的根目录。"""
    if scope == "global":
        return KB_DATA_DIR / "global"
    if scope == "project":
        pid = _safe_part(project_id, "project_id")
        d = KB_DATA_DIR / "projects" / pid
        d.mkdir(parents=True, exist_ok=True)
        return d
    if scope == "function":
        pid = _safe_part(project_id, "project_id")
        fid = _safe_part(function_id, "function_id")
        kind = _safe_part(kb_type or "general", "kb_type")
        d = KB_DATA_DIR / "functions" / pid / fid / kind
        d.mkdir(parents=True, exist_ok=True)
        return d
    raise HTTPException(400, f"非法 scope: {scope}")


def _doc_dir(
    scope: str,
    project_id: Optional[str],
    doc_id: str,
    function_id: Optional[str] = None,
    kb_type: Optional[str] = None,
) -> pathlib.Path:
    if not doc_id.startswith("doc_") or "/" in doc_id or ".." in doc_id:
        raise HTTPException(400, f"非法 doc_id: {doc_id}")
    return _scope_dir(scope, project_id, function_id, kb_type) / doc_id


def _read_meta(d: pathlib.Path) -> dict:
    fp = d / "meta.json"
    if not fp.exists():
        return {}
    try:
        return json.loads(fp.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _write_meta(d: pathlib.Path, data: dict) -> None:
    d.mkdir(parents=True, exist_ok=True)
    (d / "meta.json").write_text(
        json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def _summary_meta(meta: dict) -> dict:
    """列表页用，去掉大字段。"""
    return {
        "id": meta.get("id"),
        "scope": meta.get("scope"),
        "project_id": meta.get("project_id"),
        "function_id": meta.get("function_id"),
        "kb_type": meta.get("kb_type"),
        "filename": meta.get("filename"),
        "size": meta.get("size"),
        "format": meta.get("format"),
        "content_kind": meta.get("content_kind"),
        "tags": meta.get("tags") or [],
        "summary": meta.get("summary"),
        "chunks_count": meta.get("chunks_count", 0),
        "char_count": meta.get("char_count", 0),
        "status": meta.get("status", "ready"),  # parsing | ready | error
        "error": meta.get("error"),
        "created_at": meta.get("created_at"),
    }


def _iter_all_docs(
    scope: Optional[str] = None,
    project_id: Optional[str] = None,
    function_id: Optional[str] = None,
    kb_type: Optional[str] = None,
):
    """遍历所有文档元信息（按筛选）。"""
    if scope is None or scope == "global":
        gdir = KB_DATA_DIR / "global"
        if gdir.exists():
            for d in gdir.iterdir():
                if d.is_dir() and d.name.startswith("doc_"):
                    meta = _read_meta(d)
                    if meta:
                        yield meta
    if scope is None or scope == "function":
        fdir = KB_DATA_DIR / "functions"
        if fdir.exists():
            project_dirs = [fdir / project_id] if project_id else [p for p in fdir.iterdir() if p.is_dir()]
            for pdir in project_dirs:
                if not pdir.exists() or not pdir.is_dir():
                    continue
                function_dirs = [pdir / function_id] if function_id else [f for f in pdir.iterdir() if f.is_dir()]
                for func_dir in function_dirs:
                    if not func_dir.exists() or not func_dir.is_dir():
                        continue
                    type_dirs = [func_dir / kb_type] if kb_type else [t for t in func_dir.iterdir() if t.is_dir()]
                    for type_dir in type_dirs:
                        if not type_dir.exists() or not type_dir.is_dir():
                            continue
                        for d in type_dir.iterdir():
                            if d.is_dir() and d.name.startswith("doc_"):
                                meta = _read_meta(d)
                                if meta:
                                    yield meta
    if scope is None or scope == "project":
        pdir = KB_DATA_DIR / "projects"
        if pdir.exists():
            if project_id:
                target = pdir / project_id
                if target.exists():
                    for d in target.iterdir():
                        if d.is_dir() and d.name.startswith("doc_"):
                            meta = _read_meta(d)
                            if meta:
                                yield meta
            else:
                for proj in pdir.iterdir():
                    if not proj.is_dir():
                        continue
                    for d in proj.iterdir():
                        if d.is_dir() and d.name.startswith("doc_"):
                            meta = _read_meta(d)
                            if meta:
                                yield meta


# ============================================================
# 路由
# ============================================================

@router.get("")
def list_docs(
    scope: Optional[str] = None,
    project_id: Optional[str] = None,
    function_id: Optional[str] = None,
    kb_type: Optional[str] = None,
):
    """列出文档。
    - 不传 scope：返回所有
    - scope=global：仅全局
    - scope=project：仅指定项目（需配合 project_id）
    - scope=function：仅指定项目 + 功能 + 类型
    """
    items = []
    for meta in _iter_all_docs(scope=scope, project_id=project_id, function_id=function_id, kb_type=kb_type):
        items.append(_summary_meta(meta))
    items.sort(key=lambda x: x.get("created_at") or "", reverse=True)
    return {"docs": items}


@router.post("/upload")
async def upload_doc(
    file: UploadFile = File(...),
    scope: str = Form("global"),
    project_id: Optional[str] = Form(None),
    function_id: Optional[str] = Form(None),
    kb_type: Optional[str] = Form(None),
):
    """上传一份文档，同步解析 + 切片。"""
    if scope == "project" and not project_id:
        raise HTTPException(400, "scope=project 时必须传 project_id")
    if scope == "function":
        if not project_id:
            raise HTTPException(400, "scope=function 时必须传 project_id")
        if not function_id:
            raise HTTPException(400, "scope=function 时必须传 function_id")
        kb_type = kb_type or "general"

    raw_name = pathlib.Path(file.filename or "untitled").name
    ext = pathlib.Path(raw_name).suffix.lower()
    if ext not in SUPPORTED_EXTS:
        raise HTTPException(400, f"不支持的格式 {ext}，仅支持 {sorted(SUPPORTED_EXTS)}")

    # 读到内存检查大小（小文件直接读，大文件应该用流式，但 v0.3 暂这样）
    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(413, f"文件过大（>{MAX_FILE_SIZE // 1024 // 1024}MB），请压缩或拆分")

    doc_id = _gen_doc_id()
    ddir = _doc_dir(scope, project_id, doc_id, function_id=function_id, kb_type=kb_type)
    ddir.mkdir(parents=True, exist_ok=True)

    # 落盘原文件
    original_path = ddir / f"original{ext}"
    original_path.write_bytes(content)

    # 先写 meta（status=parsing）
    meta = {
        "id": doc_id,
        "scope": scope,
        "project_id": project_id if scope in ("project", "function") else None,
        "function_id": function_id if scope == "function" else None,
        "kb_type": kb_type if scope == "function" else None,
        "filename": raw_name,
        "size": len(content),
        "format": ext.lstrip("."),
        "content_kind": _content_kind(ext),
        "created_at": _now_iso(),
        "tags": [],
        "summary": None,
        "chunks_count": 0,
        "char_count": 0,
        "status": "parsing",
    }
    _write_meta(ddir, meta)

    # 解析 + 切片（同步，错了把 status 改成 error）
    try:
        text = load_doc(original_path)
    except LoaderError as e:
        meta["status"] = "error"
        meta["error"] = str(e)
        _write_meta(ddir, meta)
        return JSONResponse(
            status_code=400,
            content={"error": str(e), "doc_id": doc_id, "meta": _summary_meta(meta)},
        )
    except Exception as e:
        meta["status"] = "error"
        meta["error"] = f"解析失败：{e}"
        _write_meta(ddir, meta)
        return JSONResponse(
            status_code=500,
            content={"error": meta["error"], "doc_id": doc_id, "meta": _summary_meta(meta)},
        )

    text = (text or "").strip()
    if not text:
        meta["status"] = "error"
        meta["error"] = "提取出的文本为空（可能是扫描版 PDF / 加密文件）"
        _write_meta(ddir, meta)
        return JSONResponse(
            status_code=400,
            content={"error": meta["error"], "doc_id": doc_id, "meta": _summary_meta(meta)},
        )

    # 落盘 text + chunks
    (ddir / "text.txt").write_text(text, encoding="utf-8")
    chunks = chunk_text(text)
    chunks_data = [
        {"id": f"{doc_id}_c{i}", "text": c}
        for i, c in enumerate(chunks)
    ]
    (ddir / "chunks.json").write_text(
        json.dumps(chunks_data, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    # 简易摘要：取前 200 字 + chunk 数
    summary = text[:200].replace("\n", " ").strip()
    meta.update({
        "chunks_count": len(chunks),
        "char_count": len(text),
        "summary": summary + ("…" if len(text) > 200 else ""),
        "status": "ready",
    })
    _write_meta(ddir, meta)

    # 新文档进来：让索引在下次搜索时重建
    kb_index.invalidate(scope, project_id, function_id=function_id, kb_type=kb_type)

    return _summary_meta(meta)


@router.get("/{doc_id}")
def get_doc(
    doc_id: str,
    scope: str = "global",
    project_id: Optional[str] = None,
    function_id: Optional[str] = None,
    kb_type: Optional[str] = None,
):
    """取单文档元信息（含完整 summary + 前 3 个 chunk 预览）。"""
    ddir = _doc_dir(scope, project_id, doc_id, function_id=function_id, kb_type=kb_type)
    if not ddir.exists():
        raise HTTPException(404, f"文档不存在: {doc_id}")
    meta = _read_meta(ddir)
    chunks_fp = ddir / "chunks.json"
    preview = []
    if chunks_fp.exists():
        try:
            all_chunks = json.loads(chunks_fp.read_text(encoding="utf-8"))
            preview = all_chunks[:3]
        except Exception:
            pass
    return {"meta": _summary_meta(meta), "chunks_preview": preview}


@router.get("/{doc_id}/file")
def get_doc_file(
    doc_id: str,
    scope: str = "global",
    project_id: Optional[str] = None,
    function_id: Optional[str] = None,
    kb_type: Optional[str] = None,
):
    """下载/预览原文件。"""
    ddir = _doc_dir(scope, project_id, doc_id, function_id=function_id, kb_type=kb_type)
    if not ddir.exists():
        raise HTTPException(404, f"文档不存在: {doc_id}")
    # 找 original.* 文件
    matches = list(ddir.glob("original.*"))
    if not matches:
        raise HTTPException(404, "原文件不存在")
    return FileResponse(matches[0], filename=_read_meta(ddir).get("filename"))


@router.delete("/{doc_id}")
def delete_doc(
    doc_id: str,
    scope: str = "global",
    project_id: Optional[str] = None,
    function_id: Optional[str] = None,
    kb_type: Optional[str] = None,
):
    ddir = _doc_dir(scope, project_id, doc_id, function_id=function_id, kb_type=kb_type)
    if not ddir.exists():
        raise HTTPException(404, f"文档不存在: {doc_id}")
    shutil.rmtree(ddir, ignore_errors=True)
    kb_index.invalidate(scope, project_id, function_id=function_id, kb_type=kb_type)
    return {"ok": True, "id": doc_id}


@router.post("/search")
def search_kb(payload: dict):
    """检索知识库。
    payload: {query, scope?='global', project_id?, top_k?=5}
    返回：{results: [{doc_id, chunk_id, text, score, filename}]}
    """
    query = (payload.get("query") or "").strip()
    if not query:
        raise HTTPException(400, "缺少 query")
    scope = payload.get("scope") or "global"
    project_id = payload.get("project_id")
    function_id = payload.get("function_id")
    kb_type = payload.get("kb_type")
    top_k = int(payload.get("top_k") or 5)
    if top_k < 1:
        top_k = 5
    if top_k > 20:
        top_k = 20

    results = kb_index.search(
        query,
        scope=scope,
        project_id=project_id,
        function_id=function_id,
        kb_type=kb_type,
        top_k=top_k,
    )
    return {"results": results, "engine": "bm25"}


@router.post("/extract-project")
def extract_project(payload: dict):
    """从指定文档 LLM 抽取项目草案。
    payload: {doc_ids: ["doc_xxx", ...]}
    返回：auto_project.extract() 结果
    """
    doc_ids = payload.get("doc_ids") or []
    if not isinstance(doc_ids, list) or not doc_ids:
        raise HTTPException(400, "缺少 doc_ids")

    import llm_client
    from skills import auto_project

    try:
        client = llm_client.LLMClient()
        if not client.is_configured:
            raise HTTPException(503, "LLM 未配置：请在设置抽屉填入 DeepSeek API Key")
    except llm_client.LLMError as e:
        raise HTTPException(503, str(e))

    result = auto_project.extract(doc_ids, client)
    return result


@router.post("/recognize-project")
def recognize_project(payload: dict):
    """从已有项目知识库识别项目类型、场景和 M1-M25 模块编排。
    payload: {project_id}
    """
    project_id = payload.get("project_id")
    if not project_id:
        raise HTTPException(400, "缺少 project_id")
    docs = list(_iter_all_docs(scope="project", project_id=project_id))
    text = "\n\n".join(_doc_text(m) for m in docs)
    files = [fp for fp in (_original_file_for_meta(m) for m in docs) if fp]
    if not text.strip():
        raise HTTPException(400, "项目知识库暂无可读取文本，请先上传 Word/PDF/Excel/文字资料")
    return _recognition_payload_from_sources(text, files)


@router.post("/recognize-project-files")
async def recognize_project_files(files: List[UploadFile] = File(...)):
    """新建项目时直接识别待上传资料，不落全局知识库。"""
    if not files:
        raise HTTPException(400, "请上传至少一个文件")
    parts = []
    original_files = []
    tmp_root = KB_DATA_DIR / "_tmp_recognize" / f"tmp_{uuid.uuid4().hex[:10]}"
    tmp_root.mkdir(parents=True, exist_ok=True)
    try:
        for file in files:
            raw_name = pathlib.Path(file.filename or "untitled").name
            ext = pathlib.Path(raw_name).suffix.lower()
            if ext not in SUPPORTED_EXTS:
                continue
            content = await file.read()
            if len(content) > MAX_FILE_SIZE:
                continue
            fp = tmp_root / f"{uuid.uuid4().hex[:8]}{ext}"
            fp.write_bytes(content)
            original_files.append(fp)
            try:
                parts.append(load_doc(fp))
            except Exception:
                pass
        text = "\n\n".join(p for p in parts if p)
        if not text.strip():
            raise HTTPException(400, "上传资料暂未提取到可识别文本")
        return _recognition_payload_from_sources(text, original_files)
    finally:
        shutil.rmtree(tmp_root, ignore_errors=True)


@router.post("/{doc_id}/tag")
def tag_doc(
    doc_id: str,
    payload: dict,
    scope: str = "global",
    project_id: Optional[str] = None,
    function_id: Optional[str] = None,
    kb_type: Optional[str] = None,
):
    """给文档加 tag（用于标记"由哪个项目上传"）。
    payload: {tags: ["proj_xxx"]}（追加，不覆盖）
    """
    ddir = _doc_dir(scope, project_id, doc_id, function_id=function_id, kb_type=kb_type)
    if not ddir.exists():
        raise HTTPException(404, f"文档不存在: {doc_id}")
    new_tags = payload.get("tags") or []
    if not isinstance(new_tags, list):
        raise HTTPException(400, "tags 必须是数组")
    meta = _read_meta(ddir)
    cur = meta.get("tags") or []
    for t in new_tags:
        t = str(t).strip()
        if t and t not in cur:
            cur.append(t)
    meta["tags"] = cur
    _write_meta(ddir, meta)
    return _summary_meta(meta)
